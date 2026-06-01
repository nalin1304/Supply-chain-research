#!/usr/bin/env python3
"""Data ingestion script for Dalal (2022) INFORMS dataset.

Reads the supplementary DataSet.xlsx (101 Indian demand points with lat/lon/demand
for 4 product categories) and generates processed numpy arrays for the optimization
pipeline.

Also supports synthetic data generation following Dalal's methodology when the
Excel file is not available (use --synthetic flag).

Reference:
    Dalal, J. (2022). Multi-product green supply chain network design with
    location-routing and simultaneous pickup-delivery. INFORMS Journal on
    Computing, 34(1), 269-284.

Usage:
    # With real data:
    python scripts/ingest_dalal_data.py

    # With synthetic data:
    python scripts/ingest_dalal_data.py --synthetic

    # Skip OSRM distance matrix computation:
    python scripts/ingest_dalal_data.py --synthetic --skip-osrm
"""

import argparse
import hashlib
import json
import logging
import sys
import time
from pathlib import Path

import numpy as np

# Add project root to path
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from supply_chain_research.config import CFG

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger(__name__)

# Constants
N_CUSTOMERS = 101
N_PRODUCTS = 4
PRODUCT_CATEGORIES = ["Electronics", "Books", "Apparel", "Grocery"]

# Dalal demand generation ranges per preference rank
DEMAND_RANGES = {
    1: (350, 450),  # First choice product
    2: (250, 350),  # Second choice
    3: (150, 250),  # Third choice
    4: (50, 150),   # Fourth choice
}

# Output paths
DATA_DIR = PROJECT_ROOT / "data"
PROCESSED_DIR = DATA_DIR / "processed"
CACHE_DIR = DATA_DIR / "cache"
EXTERNAL_DIR = DATA_DIR / "external"

OUTPUT_LOCATIONS = PROCESSED_DIR / "dalal_customer_locations.npy"
OUTPUT_DEMAND = PROCESSED_DIR / "dalal_demand.npy"
OUTPUT_DISTANCE = PROCESSED_DIR / "dalal_distance_matrix_km.npy"


def load_excel_data(filepath: Path) -> tuple:
    """Load customer locations and demand from Dalal (2022) DataSet.xlsx.

    Expected format: Excel file with columns for customer ID, latitude,
    longitude, and demand for each of 4 product categories.

    Returns:
        Tuple of (locations, demand) numpy arrays.
    """
    try:
        import openpyxl
    except ImportError:
        logger.error(
            "openpyxl is required to read Excel files. "
            "Install with: pip install openpyxl"
        )
        raise

    logger.info(f"Reading Excel file: {filepath}")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # Try common sheet names
    sheet = None
    for name in wb.sheetnames:
        if "customer" in name.lower() or "demand" in name.lower() or "data" in name.lower():
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb.active

    logger.info(f"Using sheet: {sheet.title}")

    rows = list(sheet.iter_rows(min_row=2, values_only=True))  # Skip header

    locations = []
    demand = []

    for row in rows:
        if row[0] is None:
            continue
        # Expected columns: ID, Lat, Lon, Demand1, Demand2, Demand3, Demand4
        # Adjust indices based on actual file structure
        try:
            lat = float(row[1])
            lon = float(row[2])
            d = [float(row[3 + i]) for i in range(N_PRODUCTS)]
            locations.append([lat, lon])
            demand.append(d)
        except (TypeError, ValueError, IndexError) as e:
            logger.warning(f"Skipping row {row[0]}: {e}")
            continue

    locations = np.array(locations, dtype=np.float64)
    demand = np.array(demand, dtype=np.float64)

    logger.info(f"Loaded {len(locations)} customer locations")
    logger.info(f"Demand shape: {demand.shape}")

    if len(locations) != N_CUSTOMERS:
        logger.warning(
            f"Expected {N_CUSTOMERS} customers, got {len(locations)}. "
            "Proceeding with available data."
        )

    return locations, demand


def generate_synthetic_data(seed: int = 42) -> tuple:
    """Generate synthetic data following Dalal (2022) methodology.

    Generates 101 customer locations spread across India using the 20 cities
    from config.py as cluster centers, with Gaussian perturbation.

    Demand is generated per Dalal's algorithm:
    - Each customer has a random preference ordering over 4 products
    - First choice: uniform(350, 450)
    - Second choice: uniform(250, 350)
    - Third choice: uniform(150, 250)
    - Fourth choice: uniform(50, 150)

    Returns:
        Tuple of (locations, demand, warehouse_locations) numpy arrays.
    """
    rng = np.random.default_rng(seed)
    cities = CFG.network.cities

    logger.info("Generating synthetic data following Dalal (2022) methodology")
    logger.info(f"  Customers: {N_CUSTOMERS}")
    logger.info(f"  Products: {PRODUCT_CATEGORIES}")
    logger.info(f"  Cities used as cluster centers: {len(cities)}")

    # Generate customer locations by assigning each to a random city
    # and adding Gaussian perturbation
    locations = np.zeros((N_CUSTOMERS, 2))
    city_assignments = rng.integers(0, len(cities), size=N_CUSTOMERS)

    for i in range(N_CUSTOMERS):
        city_name, city_lat, city_lon = cities[city_assignments[i]]
        # Gaussian perturbation (std from config, ~0.8 degrees ≈ 89 km)
        std = CFG.network.customer_location_std
        lat = city_lat + rng.normal(0, std)
        lon = city_lon + rng.normal(0, std)

        # Clip to India bounds
        lat = np.clip(lat, *CFG.network.india_lat_bounds)
        lon = np.clip(lon, *CFG.network.india_lon_bounds)

        locations[i] = [lat, lon]

    # Generate demand per Dalal's algorithm
    demand = np.zeros((N_CUSTOMERS, N_PRODUCTS))
    for i in range(N_CUSTOMERS):
        # Random preference ordering over 4 products
        pref_order = rng.permutation(N_PRODUCTS)
        for rank, product_idx in enumerate(pref_order, start=1):
            low, high = DEMAND_RANGES[rank]
            demand[i, product_idx] = rng.uniform(low, high)

    # Select 3-5 warehouse locations as cluster centroids using k-means
    n_warehouses = rng.integers(3, 6)  # 3 to 5
    logger.info(f"  Warehouses (cluster centroids): {n_warehouses}")

    # Simple k-means to find warehouse locations
    warehouse_locations = _kmeans_centroids(locations, n_warehouses, rng)

    logger.info(f"  Location bounds: lat [{locations[:, 0].min():.2f}, {locations[:, 0].max():.2f}]")
    logger.info(f"  Location bounds: lon [{locations[:, 1].min():.2f}, {locations[:, 1].max():.2f}]")
    logger.info(f"  Demand range: [{demand.min():.1f}, {demand.max():.1f}]")

    return locations, demand, warehouse_locations


def _kmeans_centroids(points: np.ndarray, k: int, rng: np.random.Generator,
                      max_iter: int = 100) -> np.ndarray:
    """Simple k-means to find cluster centroids for warehouse placement."""
    n = len(points)
    # Initialize with random points
    indices = rng.choice(n, size=k, replace=False)
    centroids = points[indices].copy()

    for _ in range(max_iter):
        # Assign points to nearest centroid
        dists = np.linalg.norm(points[:, None, :] - centroids[None, :, :], axis=2)
        assignments = np.argmin(dists, axis=1)

        # Update centroids
        new_centroids = np.zeros_like(centroids)
        for j in range(k):
            mask = assignments == j
            if mask.any():
                new_centroids[j] = points[mask].mean(axis=0)
            else:
                new_centroids[j] = centroids[j]

        if np.allclose(centroids, new_centroids, atol=1e-6):
            break
        centroids = new_centroids

    return centroids


def compute_osrm_distance_matrix(
    locations: np.ndarray,
    cache_dir: Path = CACHE_DIR,
    base_url: str = None,
) -> np.ndarray:
    """Compute distance matrix using OSRM table service with caching.

    Uses the OSRM table API to compute driving distances between all pairs
    of locations. Results are cached to disk based on a hash of the input
    coordinates.

    Args:
        locations: (N, 2) array of (latitude, longitude) coordinates.
        cache_dir: Directory for caching distance matrices.
        base_url: OSRM server URL. Defaults to config value.

    Returns:
        (N, N) distance matrix in kilometers.
    """
    import requests

    if base_url is None:
        base_url = CFG.network.osrm_base_url

    n = len(locations)
    logger.info(f"Computing {n}x{n} OSRM distance matrix...")

    # Check cache
    cache_dir.mkdir(parents=True, exist_ok=True)
    coord_hash = hashlib.md5(locations.tobytes()).hexdigest()[:12]
    cache_file = cache_dir / f"osrm_dist_{coord_hash}_{n}x{n}.npy"
    meta_file = cache_dir / f"osrm_dist_{coord_hash}_{n}x{n}.json"

    if cache_file.exists():
        logger.info(f"Loading cached distance matrix: {cache_file}")
        return np.load(cache_file)

    # OSRM table API expects coordinates as lon,lat (note: reversed from lat,lon)
    coords_str = ";".join(
        f"{locations[i, 1]:.6f},{locations[i, 0]:.6f}" for i in range(n)
    )

    # OSRM has a limit on URL length; batch if needed
    batch_size = CFG.network.osrm_batch_size
    if n <= batch_size:
        distance_matrix = _osrm_table_request(
            base_url, coords_str, n, locations
        )
    else:
        distance_matrix = _osrm_batched_table(
            base_url, locations, batch_size
        )

    # Save to cache
    np.save(cache_file, distance_matrix)
    meta = {
        "n_points": n,
        "coord_hash": coord_hash,
        "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
        "source": "OSRM",
    }
    meta_file.write_text(json.dumps(meta, indent=2))
    logger.info(f"Cached distance matrix to: {cache_file}")

    return distance_matrix


def _osrm_table_request(
    base_url: str, coords_str: str, n: int, locations: np.ndarray
) -> np.ndarray:
    """Make a single OSRM table API request."""
    import requests

    url = f"{base_url}/table/v1/driving/{coords_str}"
    params = {"annotations": "distance"}

    for attempt in range(CFG.network.osrm_retry_max):
        try:
            resp = requests.get(url, params=params, timeout=60)
            resp.raise_for_status()
            data = resp.json()

            if data.get("code") != "Ok":
                raise RuntimeError(f"OSRM error: {data.get('message', 'Unknown')}")

            # Convert from meters to kilometers
            distances = np.array(data["distances"], dtype=np.float64) / 1000.0
            logger.info(
                f"OSRM matrix computed: "
                f"min={distances[distances > 0].min():.1f} km, "
                f"max={distances.max():.1f} km, "
                f"mean={distances[distances > 0].mean():.1f} km"
            )
            return distances

        except (requests.RequestException, RuntimeError) as e:
            wait = CFG.network.osrm_retry_backoff ** attempt
            logger.warning(
                f"OSRM request failed (attempt {attempt + 1}): {e}. "
                f"Retrying in {wait:.1f}s..."
            )
            time.sleep(wait)

    raise RuntimeError(
        f"OSRM table request failed after {CFG.network.osrm_retry_max} attempts. "
        "Falling back to Haversine distances."
    )


def _osrm_batched_table(
    base_url: str, locations: np.ndarray, batch_size: int
) -> np.ndarray:
    """Compute distance matrix in batches for large point sets."""
    n = len(locations)
    distance_matrix = np.zeros((n, n), dtype=np.float64)

    # For large matrices, compute row-by-row batches
    for i in range(0, n, batch_size):
        i_end = min(i + batch_size, n)
        source_indices = list(range(i, i_end))

        # Build coordinate string with all points
        coords_str = ";".join(
            f"{locations[j, 1]:.6f},{locations[j, 0]:.6f}" for j in range(n)
        )

        sources_param = ";".join(str(s) for s in source_indices)
        url = f"{base_url}/table/v1/driving/{coords_str}"
        params = {
            "annotations": "distance",
            "sources": sources_param,
        }

        import requests
        try:
            resp = requests.get(url, params=params, timeout=120)
            resp.raise_for_status()
            data = resp.json()

            if data.get("code") == "Ok":
                batch_distances = np.array(data["distances"], dtype=np.float64) / 1000.0
                distance_matrix[i:i_end, :] = batch_distances
            else:
                logger.warning(f"OSRM batch failed, using Haversine fallback for rows {i}:{i_end}")
                distance_matrix[i:i_end, :] = _haversine_matrix(
                    locations[i:i_end], locations
                )
        except requests.RequestException as e:
            logger.warning(f"OSRM batch request failed: {e}. Using Haversine fallback.")
            distance_matrix[i:i_end, :] = _haversine_matrix(
                locations[i:i_end], locations
            )

        # Rate limiting
        time.sleep(1.0)

    return distance_matrix


def _haversine_matrix(sources: np.ndarray, targets: np.ndarray) -> np.ndarray:
    """Compute Haversine distance matrix as fallback (km).

    Applies a 1.3x road circuity factor to approximate driving distances.
    """
    EARTH_RADIUS_KM = 6371.0
    CIRCUITY_FACTOR = 1.3  # Approximate road/straight-line ratio for India

    lat1 = np.radians(sources[:, 0])[:, None]
    lon1 = np.radians(sources[:, 1])[:, None]
    lat2 = np.radians(targets[:, 0])[None, :]
    lon2 = np.radians(targets[:, 1])[None, :]

    dlat = lat2 - lat1
    dlon = lon2 - lon1

    a = np.sin(dlat / 2) ** 2 + np.cos(lat1) * np.cos(lat2) * np.sin(dlon / 2) ** 2
    c = 2 * np.arcsin(np.sqrt(a))

    return EARTH_RADIUS_KM * c * CIRCUITY_FACTOR


def compute_haversine_distance_matrix(locations: np.ndarray) -> np.ndarray:
    """Compute full Haversine distance matrix with circuity factor.

    Used as fallback when OSRM is unavailable.
    """
    return _haversine_matrix(locations, locations)


def main():
    """Main entry point for data ingestion."""
    parser = argparse.ArgumentParser(
        description="Ingest Dalal (2022) dataset or generate synthetic equivalent"
    )
    parser.add_argument(
        "--synthetic",
        action="store_true",
        help="Generate synthetic data following Dalal methodology",
    )
    parser.add_argument(
        "--skip-osrm",
        action="store_true",
        help="Skip OSRM distance computation (use Haversine fallback)",
    )
    parser.add_argument(
        "--seed",
        type=int,
        default=42,
        help="Random seed for synthetic data generation (default: 42)",
    )
    parser.add_argument(
        "--input",
        type=str,
        default=str(EXTERNAL_DIR / "DataSet.xlsx"),
        help="Path to Dalal DataSet.xlsx file",
    )
    args = parser.parse_args()

    # Ensure output directories exist
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    CACHE_DIR.mkdir(parents=True, exist_ok=True)

    # Load or generate data
    if args.synthetic:
        locations, demand, warehouse_locs = generate_synthetic_data(seed=args.seed)
        logger.info(f"Generated {len(warehouse_locs)} warehouse locations")
    else:
        input_path = Path(args.input)
        if not input_path.exists():
            logger.error(
                f"Input file not found: {input_path}\n"
                f"Please place the Dalal (2022) DataSet.xlsx at:\n"
                f"  {EXTERNAL_DIR / 'DataSet.xlsx'}\n\n"
                f"Or use --synthetic to generate equivalent data."
            )
            sys.exit(1)
        locations, demand = load_excel_data(input_path)

    # Save locations and demand
    np.save(OUTPUT_LOCATIONS, locations)
    logger.info(f"Saved customer locations: {OUTPUT_LOCATIONS} (shape: {locations.shape})")

    np.save(OUTPUT_DEMAND, demand)
    logger.info(f"Saved demand matrix: {OUTPUT_DEMAND} (shape: {demand.shape})")

    # Compute distance matrix
    if args.skip_osrm:
        logger.info("Computing Haversine distance matrix (OSRM skipped)")
        distance_matrix = compute_haversine_distance_matrix(locations)
    else:
        try:
            distance_matrix = compute_osrm_distance_matrix(locations)
        except RuntimeError as e:
            logger.warning(f"OSRM failed: {e}")
            logger.info("Falling back to Haversine distance matrix")
            distance_matrix = compute_haversine_distance_matrix(locations)

    np.save(OUTPUT_DISTANCE, distance_matrix)
    logger.info(
        f"Saved distance matrix: {OUTPUT_DISTANCE} "
        f"(shape: {distance_matrix.shape})"
    )

    # Print summary statistics
    nonzero = distance_matrix[distance_matrix > 0]
    logger.info("Distance matrix statistics:")
    logger.info(f"  Min (non-zero): {nonzero.min():.1f} km")
    logger.info(f"  Max: {nonzero.max():.1f} km")
    logger.info(f"  Mean: {nonzero.mean():.1f} km")
    logger.info(f"  Std: {nonzero.std():.1f} km")

    logger.info("Data ingestion complete.")


if __name__ == "__main__":
    main()
